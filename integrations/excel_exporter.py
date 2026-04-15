from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


@dataclass
class ExcelExportResult:
    """
    Resultado de exportación técnica a Excel.
    """
    success: bool
    file_path: str = ""
    sheet_name: str = ""
    written_rows: int = 0
    error_message: str = ""

    def to_dict(self) -> dict:
        return {
            "success": self.success,
            "file_path": self.file_path,
            "sheet_name": self.sheet_name,
            "written_rows": self.written_rows,
            "error_message": self.error_message,
        }


def create_workbook() -> Workbook:
    """
    Crea un workbook nuevo.
    """
    return Workbook()


def get_or_create_worksheet(workbook: Workbook, sheet_name: str):
    """
    Obtiene una hoja o la crea si no existe.
    """
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    # Si es el libro recién creado y la hoja activa no tiene uso,
    # se puede reutilizar la hoja por defecto.
    if workbook.active and workbook.active.title == "Sheet" and workbook.active.max_row == 1 and workbook.active.max_column == 1:
        ws = workbook.active
        ws.title = sheet_name
        return ws

    return workbook.create_sheet(title=sheet_name)


def write_headers(worksheet, headers: list[str]) -> None:
    """
    Escribe encabezados en la primera fila.
    """
    if not headers:
        return

    worksheet.delete_rows(1, 1)
    worksheet.insert_rows(1)

    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)


def append_rows(worksheet, rows: list[dict], headers: list[str] | None = None) -> int:
    """
    Agrega filas a una hoja.
    """
    if not rows:
        return 0

    if headers is None:
        headers = list(rows[0].keys())

    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    return len(rows)


def auto_adjust_columns(worksheet, max_width: int = 50) -> None:
    """
    Ajusta ancho de columnas según contenido.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def save_workbook(workbook: Workbook, file_path: str | Path) -> None:
    """
    Guarda el workbook en disco.
    """
    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)


def export_rows_to_excel_file(
    rows: list[dict],
    file_path: str | Path,
    sheet_name: str = "Resultados",
    replace_sheet: bool = True,
    adjust_columns: bool = True,
) -> ExcelExportResult:
    """
    Exporta filas tabulares a un archivo Excel.
    """
    try:
        file_path = Path(file_path)

        workbook = create_workbook()
        worksheet = get_or_create_worksheet(workbook, sheet_name)

        if not rows:
            save_workbook(workbook, file_path)
            return ExcelExportResult(
                success=True,
                file_path=str(file_path),
                sheet_name=sheet_name,
                written_rows=0,
                error_message="",
            )

        headers = list(rows[0].keys())

        if replace_sheet:
            # limpiar contenido existente
            worksheet.delete_rows(1, worksheet.max_row if worksheet.max_row > 0 else 1)

        write_headers(worksheet, headers)
        written_rows = append_rows(worksheet, rows, headers=headers)

        if adjust_columns:
            auto_adjust_columns(worksheet)

        save_workbook(workbook, file_path)

        return ExcelExportResult(
            success=True,
            file_path=str(file_path),
            sheet_name=sheet_name,
            written_rows=written_rows,
            error_message="",
        )

    except Exception as exc:
        return ExcelExportResult(
            success=False,
            file_path=str(file_path),
            sheet_name=sheet_name,
            written_rows=0,
            error_message=f"Error al exportar archivo Excel: {exc}",
        )


def append_rows_to_existing_excel_file(
    rows: list[dict],
    file_path: str | Path,
    sheet_name: str = "Resultados",
    adjust_columns: bool = True,
) -> ExcelExportResult:
    """
    Agrega filas a un archivo Excel existente o lo crea si no existe.
    """
    try:
        file_path = Path(file_path)

        if file_path.exists():
            workbook = load_workbook(file_path)
        else:
            workbook = create_workbook()

        worksheet = get_or_create_worksheet(workbook, sheet_name)

        if not rows:
            save_workbook(workbook, file_path)
            return ExcelExportResult(
                success=True,
                file_path=str(file_path),
                sheet_name=sheet_name,
                written_rows=0,
                error_message="",
            )

        headers = list(rows[0].keys())

        # si la hoja está vacía, crear encabezados
        first_row_values = [worksheet.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]
        if not any(first_row_values):
            write_headers(worksheet, headers)

        written_rows = append_rows(worksheet, rows, headers=headers)

        if adjust_columns:
            auto_adjust_columns(worksheet)

        save_workbook(workbook, file_path)

        return ExcelExportResult(
            success=True,
            file_path=str(file_path),
            sheet_name=sheet_name,
            written_rows=written_rows,
            error_message="",
        )

    except Exception as exc:
        return ExcelExportResult(
            success=False,
            file_path=str(file_path),
            sheet_name=sheet_name,
            written_rows=0,
            error_message=f"Error al agregar filas a Excel: {exc}",
        )