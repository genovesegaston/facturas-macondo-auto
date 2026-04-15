from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models.batch_result import BatchResult
from models.invoice_data import InvoiceData
from models.validation_result import ValidationResult


@dataclass
class ExportResult:
    """
    Resultado de exportación.
    """
    success: bool = False
    export_path: str = ""
    exported_rows: int = 0
    error_message: str = ""

    def to_dict(self) -> dict:
        return {
            "success": self.success,
            "export_path": self.export_path,
            "exported_rows": self.exported_rows,
            "error_message": self.error_message,
        }


def build_export_row(invoice: InvoiceData, validation: ValidationResult | None = None) -> dict:
    """
    Construye una fila exportable combinando InvoiceData y ValidationResult.
    """
    row = invoice.to_row_dict()

    if validation is not None:
        row.update({
            "validation_is_valid": validation.is_valid,
            "validation_missing_fields": ", ".join(validation.missing_fields),
            "validation_receiver_valid": validation.receiver_valid,
            "validation_amounts_valid": validation.amounts_valid,
            "validation_rounding_difference": validation.rounding_difference,
            "validation_duplicate_detected": validation.duplicate_detected,
            "validation_duplicate_reference": validation.duplicate_reference,
            "validation_duplicate_key": validation.duplicate_key,
            "validation_confidence_level": validation.confidence_level,
            "validation_warnings": " | ".join(validation.warnings),
            "validation_errors": " | ".join(validation.errors),
            "validation_system_notes": validation.system_notes,
        })

    return row


def build_export_rows(
    invoices: list[InvoiceData],
    validations: list[ValidationResult] | None = None,
) -> list[dict]:
    """
    Construye todas las filas exportables del lote.
    """
    rows: list[dict] = []

    if validations is None:
        validations = [None] * len(invoices)

    for invoice, validation in zip(invoices, validations):
        rows.append(build_export_row(invoice, validation))

    return rows


def auto_adjust_columns(ws) -> None:
    """
    Ajusta ancho de columnas según contenido.
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def export_rows_to_excel(
    rows: list[dict],
    export_path: str | Path,
    sheet_name: str = "Resultados",
) -> ExportResult:
    """
    Exporta filas tabulares a Excel.
    """
    try:
        export_path = Path(export_path)
        export_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not rows:
            wb.save(export_path)
            return ExportResult(
                success=True,
                export_path=str(export_path),
                exported_rows=0,
                error_message="",
            )

        headers = list(rows[0].keys())
        ws.append(headers)

        for row in rows:
            ws.append([row.get(header, "") for header in headers])

        auto_adjust_columns(ws)
        wb.save(export_path)

        return ExportResult(
            success=True,
            export_path=str(export_path),
            exported_rows=len(rows),
            error_message="",
        )

    except Exception as exc:
        return ExportResult(
            success=False,
            export_path=str(export_path),
            exported_rows=0,
            error_message=f"Error al exportar a Excel: {exc}",
        )


def export_batch_to_excel(
    batch: BatchResult,
    export_path: str | Path,
    sheet_name: str = "Resultados",
) -> ExportResult:
    """
    Exporta un BatchResult completo a Excel.
    """
    rows = build_export_rows(
        invoices=batch.invoices,
        validations=batch.validation_results,
    )

    result = export_rows_to_excel(
        rows=rows,
        export_path=export_path,
        sheet_name=sheet_name,
    )

    if result.success:
        batch.export_path = result.export_path

    return result